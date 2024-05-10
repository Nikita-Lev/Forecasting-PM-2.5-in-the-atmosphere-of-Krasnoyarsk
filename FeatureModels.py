import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from tqdm import tqdm

from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score, mean_absolute_percentage_error


def Metrics(y, y_pred):
    '''Вычисление метрик качества прогнозов
    '''
    
    mse = round(mean_squared_error(y, y_pred), 2)
    mae = round(mean_absolute_error(y, y_pred), 2)
    mape = round(mean_absolute_percentage_error(y, y_pred), 2)
    R2 = round(r2_score(y, y_pred), 2)
    
    return [mse, mae, mape, R2]


def MetricsDF(model, x_train, y_train, x_valid = [], y_valid = [], tableName = 'Errors', metricsCV = [], testUsed = False, x_test = [], y_test = [], plot = False, savePath = '', sheet_name = 'По годам CV'):
    '''Формирование таблицы метрик на тренировочной и тестовой выборке
    '''
    
    
    resDf = pd.DataFrame(index = ['MSE', 'MAE', 'MAPE', 'R2'],
                         data = {'Train': Metrics(y_train, model.predict(x_train))}
                        )
    
    if len(x_valid):
        resDf['Valid'] = Metrics(y_valid, model.predict(x_valid))
        
    
    if len(metricsCV):
        resDf['CV'] = metricsCV
    
    if testUsed:
        resDf['Test'] = Metrics(y_test, model.predict(x_test))
                        
    resDf.index.name = tableName
    
    
    display(resDf)
    
    
    if plot and len(x_valid):
        x_valid.sort_index(inplace = True)
        y_valid.sort_index(inplace = True)
        CompareGraph(y_valid.index, y_valid, model.predict(x_valid), 'Истинные значения', 'Прогнозируемые значения', 'Сравнение прогноза с истинными данными')
    
    if plot and testUsed:
        x_test.sort_index(inplace = True)
        y_test.sort_index(inplace = True)
        CompareGraph(y_test.index, y_test, model.predict(x_test), 'Истинные значения', 'Прогнозируемые значения', 'Сравнение прогноза с истинными данными. Test')
    
    # Сохранение в файл
    if savePath:
        
        # Номер последней строки в файле
        numRow = 0
        wb = load_workbook(savePath)
        if sheet_name in wb.sheetnames:
            numRow = wb[sheet_name].max_row + 2
        
        with pd.ExcelWriter(savePath, engine="openpyxl", mode = 'a', if_sheet_exists='overlay') as writer:
            resDf.to_excel(writer, sheet_name = sheet_name, startrow = numRow)


def PlotGraph(x, y, xlabel, ylabel, title):
    '''Отрисовка графика функции
    '''
    
    plt.figure(figsize = (8, 3))
    plt.plot(x, y)

    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)

    plt.grid()
    plt.show()
            
        
def CompareGraph(x, y1, y2, l1, l2, title):
    '''Сравнение графиков двух списков значений
    '''
    
    plt.figure(figsize = (10, 4))
    plt.plot(x, y1, label = l1)
    plt.plot(x, y2, label = l2)

    plt.xlabel('time')
    plt.ylabel('PM 2.5, мкг/м³')
    plt.title(title)
    
    plt.legend()
    plt.grid()
    
    plt.show()
    
    
def splitDataBySeason(df, season):
    '''Разделение данных по сезонам
    '''
    
    # Зима
    if season == 'winters':
        win19 = df[: '2019-03-01 00:00:00']
        win19_20 = df['2019-11-27 00:00:00' : '2020-02-27 00:00:00']
        win20_21 = df['2020-11-27 00:00:00' : '2021-02-25 00:00:00']
        win21_22 = df['2021-12-15 00:00:00' : '2022-02-22 00:00:00']
        win22_23 = df['2022-12-13 00:00:00' : '2023-02-22 00:00:00']
        win23_24 = df['2023-11-28 00:00:00' : '2024-02-20 00:00:00']
        return pd.concat([win19, win19_20, win20_21, win21_22, win22_23, win23_24])
    
    # Весна
    if season == 'springs':
        spr19 = pd.concat([df['2019-03-10' : '2019-05-01'], df['2019-11-01' : '2019-11-22']])
        spr20 = pd.concat([df['2020-03-01' : '2020-05-01'], df['2020-11-01' : '2020-11-10']])
        spr21 = pd.concat([df['2021-03-10' : '2021-05-01'], df['2021-11-01' : '2021-11-25']])
        spr22 = pd.concat([df['2022-03-01' : '2022-05-01'], df['2022-11-01' : '2022-11-25']])
        spr23 = pd.concat([df['2023-03-01' : '2023-05-01'], df['2023-11-01' : '2023-11-27']])
        spr24 = df['2024-03-01' :]

        return pd.concat([spr19, spr20, spr21, spr22, spr23, spr24])

    # Лето
    if season == 'summers':
        sum19 = df['2019-05-01' : '2019-07-12']
        sum20 = df['2020-05-01' : '2020-08-01']
        sum21 = df['2021-05-01' : '2021-08-01']
        sum22 = df['2022-05-01' : '2022-08-01']
        sum23 = df['2023-05-03' : '2023-08-05']

        return pd.concat([sum19, sum20, sum21, sum22, sum23])
    
    # Осень
    if season == 'autumns':
        aut19 = df['2019-08-21' : '2019-10-16']
        aut20 = df['2020-08-17' : '2020-10-30']
        aut21 = df['2021-08-15' : '2021-11-01']
        aut22 = df['2022-08-01' : '2022-10-14']
        aut23 = df['2023-08-10' : '2023-10-25']

        return pd.concat([aut19, aut20, aut21, aut22, aut23])
    

def ModelProcessing(model, sensor, fill, district, begin, end, seasonModel = '', valid_size = 0.3, season_test = '', season_forecast = '', params = None, gridSearch = True, paramDependencies = [], plotRes = False, featImp = False,
                       path = '../data/', savePath = '', sheet_name = 'Results'):
    '''Обучение, подбор параметров, валидация, прогнозирование для выбранного класса моделей
    '''
    
    # Данные pm 
    df = pd.read_csv(f"{path}pm25_{sensor}{fill}.csv", sep = ';', index_col = ['Date'], parse_dates = ['Date'], usecols = ['Date', district])
    df.rename(columns = {district : 'pm'}, inplace = True)

        
    
    features = ['Pressure', 'Temperature', 'Wet', 'Wind_dir', 'Wind_speed']
    for feat in features:
        df[feat] = pd.read_csv(f"{path}Features{fill}/{feat}_{sensor}.csv", sep = ';', index_col = ['Date'], parse_dates = ['Date'], usecols = ['Date', district])

        # Проверка на пропуски
        nans = pd.isnull(df[begin : end].values).sum()
             
        if nans > 0:
            print(f'Пропуски в {feat} :', nans)
 
    
    # Добавление инверсий
    inversions = pd.read_csv(f"{path}Inversions.csv", sep = ';', index_col = ['Date'], parse_dates = ['Date'])
        
    df = inversions.join(df)    
    
    '''
    testUsed = True
    
    #idx = [f'2023-02-{i}' for i in range(18, 23)][:-2]
    #idx = [f'2023-03-{i}' for i in range(23, 28)][:-2]
    #idx = [f'2022-07-{i}' for i in range(28, 32)][:-1]
    #idx = [f'2022-10-{i}' for i in range(10, 15)][:-2]
    
    x_test = df.loc[idx].drop(['pm'], axis = 1)
    y_test = df.loc[idx]['pm']
    '''
    
    df.dropna(inplace = True)    
    
    testUsed = False
    
    # Исключение тестовой выборки и выборки прогнозирования для каждого сезона
    test_index = pd.read_csv(f"../data/test_index.csv", sep = ';', dayfirst = True, parse_dates = [0, 1, 2, 3])
    forecast_index = pd.read_csv(f"../data/forecast_index.csv", sep = ';', dayfirst = True, parse_dates = [0, 1, 2, 3])
    for col in ['winters', 'springs', 'summers', 'autumns']:
        idx_test = list(set(test_index[col].dropna()) & set(df.index))
       
        if col == season_test:
            testUsed = True
            x_test = df.loc[idx_test].drop(['pm'], axis = 1)
            y_test = df.loc[idx_test]['pm']
            
        idx_forecast = list(set(forecast_index[col].dropna()) & set(df.index))
        
        if col == season_forecast:
            x_forecast = df.loc[idx_forecast].drop(['pm'], axis = 1)
            y_forecast = df.loc[idx_forecast]['pm']
        
        df.drop(list(set(idx_test) | set(idx_forecast)), inplace = True)
    
    
    df = df[begin : end]
    
    
    tableName = f'{min(df.index.date)} — {max(df.index.date)}'
    
    
    # Срез для выбранного сезона
    if seasonModel:
        df = splitDataBySeason(df, seasonModel)
    
    # Разбиение данных
    x = df.drop(['pm'], axis = 1)
    y = df['pm']

    # Допольнительная колонка ошибок по кросс-валидации
    metricsCV = []
    
    # Поиск по сетке (оценка качества по кросс-валидации)
    if gridSearch:
        scoring = ['neg_mean_squared_error', 'neg_mean_absolute_error', 'neg_mean_absolute_percentage_error', 'r2' ]
        model = GridSearchCV(estimator = model, param_grid = params, scoring = scoring, refit = 'r2', n_jobs = -1)
        
        # Валидация не используется
        x_train, x_valid, y_train, y_valid = x, [], y, []
        
        
        # Обучение
        model.fit(x_train, y_train)
        
        param_grid = params
        params = model.best_params_
        
        
        # Качество по кросс валидации
        scorings = pd.DataFrame(model.cv_results_)[list(map(lambda x : 'mean_test_' + x, scoring))]
        metricsCV = np.round(scorings.iloc[np.where(scorings == model.best_score_)[0][0]].values, 2)
        metricsCV[:3] = abs(metricsCV[:3])
        
        model = model.best_estimator_
        
    else:
        # train : valid выборки 1-valid_size : valid_size
        x_train, x_valid, y_train, y_valid = train_test_split(x, y, test_size = valid_size, random_state = 100)
        
        model.set_params(**params)
    
        # Обучение
        model.fit(x_train, y_train)
    
    
    tableName += f" {params}"
    
    if testUsed:
        MetricsDF(model, x_train, y_train, x_valid, y_valid, tableName, metricsCV, testUsed, x_test, y_test, plotRes, savePath, sheet_name)
    else:
        MetricsDF(model, x_train, y_train, x_valid, y_valid, tableName, metricsCV, testUsed, plotRes, savePath, sheet_name)
        
        
    if season_forecast:
        print('ПРОГНОЗЫ:')
        for i in range(0, y_forecast.shape[0], 3):
            MetricsDF(model, x_train, y_train, x_valid, y_valid, tableName, metricsCV, testUsed, x_forecast.sort_index()[i:i+3], y_forecast.sort_index()[i:i+3], plotRes, savePath, sheet_name)
    
    # Определение важности признаков
    if featImp: FeatureImportance(model)
        
        
    # Зависимость качества от параметров
    if len(paramDependencies):
        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size = 0.3, random_state = 100)
        # Название параметра
        for paramName in paramDependencies:
            loss_mse = []
            for p in tqdm(param_grid[paramName]):
                model.set_params(**{paramName: p})

                model.fit(x_train, y_train)
                loss_mse.append(mean_squared_error(model.predict(x_test), y_test))
        
            PlotGraph(param_grid[paramName], loss_mse, paramName, 'MSE', f"Зависимость MSE от {paramName}")
    

def FeatureImportance(model):
    '''Вычисление и отрисовка важности признаков
    '''
    importances = dict(sorted(zip(model.feature_names_in_, model.feature_importances_), key = lambda x: x[1], reverse = True))
    
    plt.figure(figsize = (10, 4))
    plt.bar(importances.keys(), importances.values())
    
    plt.title('Важность признаков')
    plt.xlabel('Признаки')
    plt.ylabel('Важность')
    
    plt.grid()
    plt.show()

