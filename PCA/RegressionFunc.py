### Методы для построения регрессий
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.linear_model import LinearRegression, Lasso, Ridge
from sklearn.preprocessing import PolynomialFeatures
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error, mean_absolute_percentage_error

from tqdm import tqdm

from xgboost import XGBRegressor

from openpyxl import load_workbook

# Срез количества дней для тестовой выборки
maxForecastDays = 5


def CompareGraph(x, y1, y2, l1, l2, title):
    ''' Сравнение графиков двух списков значений
    '''
    
    plt.figure(figsize = (9, 4))
    plt.plot(x, y1, label = l1)
    plt.plot(x, y2, label = l2)

    plt.xlabel('time')
    plt.ylabel('PM 2.5, мкг/м³')
    plt.title(title)
    
    plt.legend()
    plt.grid()
    
    
def PredictAndMetrics(model, X, y, rnd = 2, plot = False, modelName = ''):
    ''' Прогнозы модели и вычисление ошибок
    '''
    
    y_pred = model.predict(X)
    
    if plot:
        CompareGraph(y.index, y, y_pred, 'Истинные значения', 'Пронозируемые значения', f'Сравнение прогноза {modelName} с истинными данными')
    return list(map(lambda i: round(i, rnd), [mean_squared_error(y, y_pred), mean_absolute_error(y, y_pred), mean_absolute_percentage_error(y, y_pred), r2_score(y, y_pred), 1 - (1-r2_score(y, y_pred))*(X.shape[0] - 1) / (X.shape[0] - X.shape[1] - 1)]))

 
def splitDataBySeason(df, season):
    ''' Разделение данных по сезонам
    '''
    
    # Зима
    if season == 'winters':
        win19 = df[: '2019-03-01 00:00:00']
        win19_20 = df['2019-11-27 00:00:00' : '2020-02-27 00:00:00']
        win20_21 = df['2020-11-27 00:00:00' : '2021-02-25 00:00:00']
        win21_22 = df['2021-12-15 00:00:00' : '2022-02-22 00:00:00']
        win22_23 = df['2022-12-13 00:00:00' : '2023-02-22 00:00:00']
        win23_24 = df['2023-11-28 00:00:00' : '2024-02-20 00:00:00']
        return pd.concat([win19, win19_20, win20_21, win21_22, win22_23, win23_24])
    
    # Весна
    if season == 'springs':
        spr19 = pd.concat([df['2019-03-10' : '2019-05-01'], df['2019-11-01' : '2019-11-22']])
        spr20 = pd.concat([df['2020-03-01' : '2020-05-01'], df['2020-11-01' : '2020-11-10']])
        spr21 = pd.concat([df['2021-03-10' : '2021-05-01'], df['2021-11-01' : '2021-11-25']])
        spr22 = pd.concat([df['2022-03-01' : '2022-05-01'], df['2022-11-01' : '2022-11-25']])
        spr23 = pd.concat([df['2023-03-01' : '2023-05-01'], df['2023-11-01' : '2023-11-27']])
        spr24 = df['2024-03-01' :]

        return pd.concat([spr19, spr20, spr21, spr22, spr23, spr24])

    # Лето
    if season == 'summers':
        sum19 = df['2019-05-01' : '2019-07-12']
        sum20 = df['2020-05-01' : '2020-08-01']
        sum21 = df['2021-05-01' : '2021-08-01']
        sum22 = df['2022-05-01' : '2022-08-01']
        sum23 = df['2023-05-03' : '2023-08-05']

        return pd.concat([sum19, sum20, sum21, sum22, sum23])
    
    # Осень
    if season == 'autumns':
        aut19 = df['2019-08-21' : '2019-10-16']
        aut20 = df['2020-08-17' : '2020-10-30']
        aut21 = df['2021-08-15' : '2021-11-01']
        aut22 = df['2022-08-01' : '2022-10-14']
        aut23 = df['2023-08-10' : '2023-10-25']

        return pd.concat([aut19, aut20, aut21, aut22, aut23])


def Regression(X, y, vs = 0, modelType = 'linear', x_test = [], y_test = [], param_dependences = [], alpha = 1.0):
    '''Построение модели регрессии, подбор гиперпараметров в случае бустинга регрессий
    '''
    if vs:
        x_train, x_valid, y_train, y_valid = train_test_split(X, y, test_size = vs, random_state = 100)
    else:
        x_train, y_train = X, y
    
    #display(X.shape, x_train.shape, x_valid.shape, x_test.shape )
    
    if modelType == 'xgbReg':

        params = {
            'n_estimators' : np.linspace(100, 8000, 11).astype('int'),
            'learning_rate' : [0.05, 0.1, 0.2, 0.3, 0.5]}
        
        model = XGBRegressor(booster = 'gblinear',  random_state = 100)
                
        scoring = ['neg_mean_squared_error', 'neg_mean_absolute_error', 'neg_mean_absolute_percentage_error', 'r2' ]
        
        model = GridSearchCV(estimator = model, param_grid = params, scoring = scoring, refit = 'r2', n_jobs = -1)
        
        
        # Обучение
        model.fit(x_train, y_train)
                
        # Допольнительная колонка ошибок по кросс-валидации
        metricsCV = []
        
        # Качество по кросс валидации
        scorings = pd.DataFrame(model.cv_results_)[list(map(lambda x : 'mean_test_' + x, scoring))]
        metricsCV = np.round(scorings.iloc[np.where(scorings == model.best_score_)[0][0]].values, 2)
        metricsCV[:3] = abs(metricsCV[:3])
        
        
        # Список ошибок
        metr = []

        # Тренировочная выборка 
        metr.append(PredictAndMetrics(model, x_train, y_train))
        
        # Качество по кросс-валидации
        metr.append(np.append(metricsCV, np.nan))
        
        # Тестовая выборка
        metr.append(PredictAndMetrics(model, x_test, y_test))
        
        # Зависимость качества от параметров
        if len(param_dependences):
            xgb_top = model.best_estimator_
            
            x_train, x_test, y_train, y_test = train_test_split(X, y, test_size = 0.3, random_state = 100)
            # Название параметра
            for paramName in param_dependences:
                loss_mse = []
                for p in tqdm(params[paramName]):
                    xgb_top.set_params(**{paramName: p})

                    xgb_top.fit(x_train, y_train)
                    loss_mse.append(mean_squared_error(xgb_top.predict(x_test), y_test))

                plt.figure(figsize = (8, 3))
                plt.plot(params[paramName], loss_mse)

                plt.title(f"Зависимость MSE от {paramName}")
                plt.xlabel(paramName)
                plt.ylabel('MSE')

                plt.grid()
                plt.show()
            
        return model, metr
    
    # Создание и обучение модели
    if modelType == 'linear':
        model = LinearRegression()
    
    if modelType == 'lasso':
        model = Lasso(alpha = alpha)
        
    if modelType == 'ridge':
        model = Ridge(alpha = alpha)
        
    model.fit(x_train, y_train)
    
    # Список ошибок
    metr = []
    
    # Тренировочная выборка 
    metr.append(PredictAndMetrics(model, x_train, y_train))
    
    # Валидационная выборка
    if vs:
        metr.append(PredictAndMetrics(model, x_valid, y_valid))
    
    # Тестовая выборка
    metr.append(PredictAndMetrics(model, x_test, y_test))
    
    return model, metr


def PolynomFeat(X, degree):
    ''' Возведение признаков в степень
    '''
    
    # Возведение признаков в степень
    pol = PolynomialFeatures(degree = degree)
    return pol.fit_transform(X)


def BuildModels(data, modelType = 'linReg', testSeason = '', param_dependences = [], forecast = False, plot_forecast = False):
    ''' Построение различных моделей регрессии
    '''
    
    vs = 0.2
    ts = 0.1
    
    
    # Формирование тестовой выборки для данного сезона
    x_test, y_test = [], []
    if testSeason:
        test_index = pd.read_csv(f"../data/test_index.csv", sep = ';', dayfirst = True, parse_dates = [0, 1, 2, 3])
        for col in test_index.columns:
            idx = list(set(test_index[col].dropna()) & set(data.index))


            if col == testSeason:
                x_test = data.loc[idx].drop(['pm'], axis = 1)
                y_test = data.loc[idx]['pm']

            data.drop(idx, inplace = True)
    # Целевая переменная и признаки
    y = data['pm']
    x = data.drop(['pm'], axis = 1)
    
    metricNames = ['MSE', 'MAE', 'MAPE', 'R2', 'R2_adj']

    # Бустинг регрессий
    if modelType == 'xgbReg':
        vs = 0
        xgbRes = pd.DataFrame(index = metricNames)
        
        # Построение моделей и вычисление ошибок
        model, metr = Regression(x, y, vs, modelType, x_test, y_test, param_dependences = param_dependences)
        
        pd.DataFrame(index = ['1',' 2', '3'], data = {'Train' : [4, 5, 6]})
        xgbRes['Train'] = metr[0]
        xgbRes['CV'] = metr[1]
        
        if testSeason: xgbRes['Test'] = metr[2]
        
        if forecast:
            xgbRes['Forecast'] = PredictAndMetrics(model, X_forecast, y_forecast, plot = plot_forecast)
        
        xgbRes.index.name = str(model.best_params_)
        
        display(xgbRes)
        
        return [xgbRes]

    
    linTypes = ['linear', 'lasso', 'ridge']
    polDegr = [2, 3]
    
    train_res = pd.DataFrame({'Train ' : metricNames})

    valid_res = pd.DataFrame({'Valid ' :  metricNames}) 
    
    # Используется тестовая выборка
    if testSeason: test_res = pd.DataFrame({'Test ' : metricNames})
    
    if forecast: forecast_res = pd.DataFrame({'Forecast ' : metricNames})
    

    # Линейные регрессии
    for mod in linTypes:

        # Построение моделей и вычисление ошибок
        model, metr = Regression(x, y, vs, mod, x_test, y_test)

        train_res[mod] = metr[0]
        valid_res[mod] = metr[1]
        
        if testSeason: test_res[mod] = metr[2]
        
        if forecast:
            forecast_res[mod] = PredictAndMetrics(model, X_forecast, y_forecast, plot = plot_forecast, modelName = mod)
        
    # Полиномиальные регрессии    
    for degr in polDegr:

        # Построение моделей и вычисление ошибок
        model, metr = Regression(PolynomFeat(x, degr), y, vs, x_test, y_test)
        
        polName = f'Полиномиальная {degr} степени'
        
        train_res[polName] = metr[0]
        valid_res[polName] = metr[1]
        
        if testSeason: test_res[polName] = metr[2]
            
        if forecast:
            forecast_res[polName] = PredictAndMetrics(model, PolynomFeat(X_forecast, degr), y_forecast)
            
    #print('Тренировочная выборка')
    display(train_res)
    
    #print('Валидационная выборка')
    display(valid_res)
    
    results = [train_res, valid_res]
    
    if testSeason:
        display(test_res)
        results.append(test_res)
    
    #print('Выборка прогнозирования')
    if forecast:
        display(forecast_res)
        results.append(forecast_res)
        
    return results


def featureSelect(matr, features, corrPMvalue = 0.2, corrFeatValue = 0.6):
    '''  Отбор признаков, наиболее коррелирующих с PM, исключая мультиколлинеарность
    '''
    
    # Признаки, коррелирующие с другими
    corrFeatures = []
    for feat in matr.apply(abs).sort_values(by='pm', ascending=False)[1:].index:
        # Признаки, которых нет в списке коррелирующих и которые коррелируют с PM со значением >= 0.2
        if feat not in features and feat not in corrFeatures and abs(matr[feat]['pm']) >= corrPMvalue:
            features.append(feat)
            # Исключить признаки коррелирующие с текущим
            for fcor, cor in zip(matr[feat].index[:-1], matr[feat].values[:-1]):
                
                # Если корреляция Спирмена составляет больше 0.6
                if fcor != feat and abs(cor) >= corrFeatValue:
                    corrFeatures.append(fcor)
                    

def SaveResults(savePath, fileName, df, sheetName = 'results', features = [], corr = None):
    ''' Сохранение результатов в файл
    '''
    
    # Извлечение номера последней строки в файле
    numRow = 0
    wb = load_workbook(savePath + fileName)
    
    if sheetName in wb.sheetnames:
        numRow = wb[sheetName].max_row + 4
    
    
    
    with pd.ExcelWriter(savePath + fileName, engine="openpyxl", mode = 'a', if_sheet_exists='overlay') as writer:
        # Запись признаков в файл
        if len(features):
            pd.DataFrame({'Признаки' : features, 'Корреляция Спирмена с pm' : corr}).to_excel(writer, sheet_name = sheetName, startrow = numRow, startcol = df[0].shape[1] + 1, index = False)

        for resDf in df:
            resDf.to_excel(writer, sheet_name = sheetName, startrow = numRow, index = True)

            numRow += resDf.shape[0] + 3
    
    